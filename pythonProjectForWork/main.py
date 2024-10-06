import os
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import xlrd

class ExcelFileProcessor:
    def __init__(self, master):
        self.master = master
        self.master.title("Excel File Processor")
        self.master.geometry("500x300")

        self.input_folder = ""
        self.output_file = ""

        # Создаем интерфейс
        self.create_widgets()

    def create_widgets(self):
        self.select_folder_button = tk.Button(self.master, text="Выбрать папку с Excel файлами",
                                              command=self.select_folder)
        self.select_folder_button.grid(row=0, column=0, padx=10, pady=10)

        self.folder_path_label = tk.Label(self.master, text="Путь к папке: ")
        self.folder_path_label.grid(row=0, column=1, padx=10, pady=10)

        self.select_file_button = tk.Button(self.master, text="Выбрать файл для записи", command=self.select_file)
        self.select_file_button.grid(row=1, column=0, padx=10, pady=10)

        self.file_path_label = tk.Label(self.master, text="Путь к файлу: ")
        self.file_path_label.grid(row=1, column=1, padx=10, pady=10)

        self.process_button = tk.Button(self.master, text="Запустить действия", command=self.process_files)
        self.process_button.grid(row=2, column=0, columnspan=2, pady=20)

    def select_folder(self):
        self.input_folder = filedialog.askdirectory()
        if self.input_folder:
            self.folder_path_label.config(text=f"Путь к папке: {self.input_folder}")

    def select_file(self):
        self.output_file = filedialog.askopenfilename(defaultextension=".xlsx",
                                                      filetypes=[("Excel files", "*.xlsx;*.xlsm")])
        if self.output_file:
            self.file_path_label.config(text=f"Путь к файлу: {self.output_file}")

    def process_files(self):
        if not self.input_folder or not self.output_file:
            messagebox.showwarning("Предупреждение", "Пожалуйста, выберите путь к папке и файл для записи.")
            return

        try:
            # Открываем целевой файл для записи сигналов
            wb_output = openpyxl.load_workbook(self.output_file)
            sheet_output = wb_output.active  # Работаем с активным листом
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл для записи: {e}")
            return

        files = os.listdir(self.input_folder)
        excel_files = [f for f in files if f.endswith(('.xls', '.xlsx', '.xlsm'))]

        if excel_files:
            for file in excel_files:
                file_path = os.path.join(self.input_folder, file)
                print(f"\nЧтение файла: {file}")
                try:
                    if file.endswith('.xls'):
                        # Читаем .xls файл с помощью xlrd
                        rb = xlrd.open_workbook(file_path)
                        # Получаем лист "Для программистов"
                        sheet_input = rb.sheet_by_name("Для программистов")  # Чтение конкретного листа
                        for row_idx in range(1, sheet_input.nrows):  # Пропускаем заголовок
                            signal_name = sheet_input.cell_value(row_idx, 1)  # Второй столбец - имя сигнала

                            # Проверка, что сигнал не пустой
                            if signal_name:
                                # Поиск первой незаполненной строки в целевом файле
                                for idx in range(6, sheet_output.max_row + 1):  # Начинаем с 6-й строки
                                    if sheet_output.cell(row=idx,
                                                         column=6).value is None:  # Проверка пустой ячейки в 6-м столбце
                                        sheet_output.cell(row=idx,
                                                          column=6).value = signal_name  # Записываем сигнал в 6-й столбец
                                        print(f"Добавлен сигнал {signal_name} в строке {idx}")
                                        break  # Переходим к следующему сигналу после записи
                                    else:
                                        print(f"Строка {idx} уже заполнена, пропускаем...")

                    elif file.endswith(('.xlsx', '.xlsm')):
                        # Читаем .xlsx файл с помощью openpyxl
                        wb_input = openpyxl.load_workbook(file_path)
                        # Получаем лист "Для программистов"
                        sheet_input = wb_input["Для программистов"]  # Чтение конкретного листа
                        for row in sheet_input.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
                            signal_name = row[1]  # Второй столбец - имя сигнала

                            # Проверка, что сигнал не пустой
                            if signal_name:
                                # Поиск первой незаполненной строки в целевом файле
                                for idx in range(6, sheet_output.max_row + 1):  # Начинаем с 6-й строки
                                    if sheet_output.cell(row=idx,
                                                         column=6).value is None:  # Проверка пустой ячейки в 6-м столбце
                                        sheet_output.cell(row=idx,
                                                          column=6).value = signal_name  # Записываем сигнал в 6-й столбец
                                        print(f"Добавлен сигнал {signal_name} в строке {idx}")
                                        break  # Переходим к следующему сигналу после записи
                                    else:
                                        print(f"Строка {idx} уже заполнена, пропускаем...")

                except Exception as e:
                    print(f"Не удалось прочитать файл {file_path}: {e}")

            # Сохранение изменений в целевом файле
            try:
                wb_output.save(self.output_file)
                messagebox.showinfo("Успех", f"Данные успешно обновлены в файле: {self.output_file}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось записать данные: {e}")
        else:
            messagebox.showinfo("Excel Files", "Нет файлов Excel в этой папке.")


# Создаем основное окно приложения
root = tk.Tk()
app = ExcelFileProcessor(root)

# Запускаем главный цикл приложения
root.mainloop()
